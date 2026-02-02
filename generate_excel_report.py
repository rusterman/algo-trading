#!/usr/bin/env python3
"""
Generate an investor-ready Excel report from DCA Strategy backtest results.
Uses config_loader and DCAStrategy directly for accurate, flexible reporting.
"""

import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)

from config_loader import load_config
from dca_strategy import DCAStrategy, load_and_prepare_data


class ExcelReportGenerator:
    """Generate a professional Excel report from backtest results."""
    
    def __init__(self, config, strategy, trades):
        self.config = config
        self.strategy = strategy
        self.trades = trades
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Backtest Report"
        self.current_row = 1
        self._setup_styles()
    
    def _setup_styles(self):
        """Define reusable cell styles."""
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        self.subheader_font = Font(bold=True, size=11, color="FFFFFF")
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        self.label_font = Font(bold=True, size=10)
        self.label_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
    
    def _set_column_widths(self):
        """Set appropriate column widths."""
        self.ws.column_dimensions['A'].width = 30
        self.ws.column_dimensions['B'].width = 20
        self.ws.column_dimensions['C'].width = 18
        self.ws.column_dimensions['D'].width = 18
        self.ws.column_dimensions['E'].width = 18
    
    def _add_section_header(self, title: str):
        """Add a section header with spacing."""
        if self.current_row > 1:
            self.current_row += 1
        
        self.ws.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = self.ws[f'A{self.current_row}']
        cell.value = title
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_align
        self.current_row += 1
    
    def _add_key_value(self, label: str, value: Any, format_type: str = 'text'):
        """Add a labeled key-value pair."""
        cell_label = self.ws[f'A{self.current_row}']
        cell_label.value = label
        cell_label.font = self.label_font
        cell_label.fill = self.label_fill
        cell_label.border = self.border
        cell_label.alignment = self.left_align
        
        cell_value = self.ws[f'B{self.current_row}']
        cell_value.value = value
        cell_value.border = self.border
        cell_value.alignment = self.right_align
        
        if format_type == 'currency':
            cell_value.number_format = '$#,##0.00'
        elif format_type == 'percentage':
            cell_value.number_format = '0.00%'
        elif format_type == 'number':
            cell_value.number_format = '0.00'
        
        self.current_row += 1
    
    def _add_table_header(self, headers: List[str]):
        """Add a table header row."""
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.border = self.border
            cell.alignment = self.center_align
        self.current_row += 1
    
    def _add_table_row(self, values: List[Any], formats: List[str] = None):
        """Add a table data row."""
        if formats is None:
            formats = ['text'] * len(values)
        
        for col_idx, (value, fmt) in enumerate(zip(values, formats), 1):
            cell = self.ws.cell(row=self.current_row, column=col_idx)
            cell.value = value
            cell.border = self.border
            cell.alignment = self.right_align if fmt != 'text' else self.center_align
            
            if fmt == 'currency':
                cell.number_format = '$#,##0.00'
            elif fmt == 'percentage':
                cell.number_format = '0.00%'
            elif fmt == 'number':
                cell.number_format = '0.00'
        
        self.current_row += 1
    
    def _extract_asset_name(self) -> str:
        """Extract asset name from CSV file path."""
        csv_file = self.config.csv_file
        return Path(csv_file).stem.split('_')[0].upper()
    
    def _calculate_monthly_stats(self) -> Dict[str, Dict[str, Any]]:
        """Calculate monthly performance statistics from trades."""
        monthly = {}
        
        for trade in self.trades:
            # Extract year-month from entry timestamp
            entry_time = trade.start_time
            month_str = entry_time.strftime('%Y-%m')
            
            if month_str not in monthly:
                monthly[month_str] = {'trades': 0, 'profit': 0.0}
            
            monthly[month_str]['trades'] += 1
            monthly[month_str]['profit'] += trade.profit_loss
        
        # Calculate profit percentage for each month
        initial_budget = self.strategy.initial_budget
        for month in monthly:
            monthly[month]['profit_pct'] = (monthly[month]['profit'] / initial_budget) * 100
        
        return monthly
    
    def _calculate_capital_stats(self) -> Dict[str, float]:
        """Calculate capital utilization statistics."""
        stats = {}
        
        if self.trades:
            invested_amounts = [t.total_invested for t in self.trades]
            initial_budget = self.strategy.initial_budget
            
            avg_capital_used = sum(invested_amounts) / len(invested_amounts)
            peak_capital_used = max(invested_amounts)
            
            stats['avg_capital_used_pct'] = (avg_capital_used / initial_budget) * 100
            stats['peak_capital_used_pct'] = (peak_capital_used / initial_budget) * 100
            stats['capital_buffer_pct'] = 100 - stats['peak_capital_used_pct']
            stats['capital_efficiency'] = sum(t.profit_loss for t in self.trades) / avg_capital_used if avg_capital_used > 0 else 0
        else:
            stats = {
                'avg_capital_used_pct': 0,
                'peak_capital_used_pct': 0,
                'capital_buffer_pct': 0,
                'capital_efficiency': 0
            }
        
        return stats
    
    def _calculate_duration_stats(self) -> Dict[str, Any]:
        """Calculate trade duration statistics."""
        stats = {}
        
        if self.trades:
            durations = [(t.end_time - t.start_time).total_seconds() / 86400.0 for t in self.trades]
            
            stats['avg_duration'] = sum(durations) / len(durations)
            stats['max_duration'] = max(durations)
            stats['min_duration'] = min(durations)
            stats['longer_than_30d'] = sum(1 for d in durations if d > 30)
        else:
            stats = {
                'avg_duration': 0,
                'max_duration': 0,
                'min_duration': 0,
                'longer_than_30d': 0
            }
        
        return stats
    
    def _calculate_drawdown_analysis(self) -> Dict[str, Dict[str, Any]]:
        """Calculate drawdown analysis from trades using actual DCA levels."""
        # Get sorted DCA levels as positive values (e.g., [6, 9, 12, 15, 18])
        dca_levels = sorted([abs(level) for level in self.config.dca_levels])
        
        # Create ranges based on actual DCA levels
        analysis = {}
        
        if dca_levels:
            # Range for trades that close before first DCA (TP before drawdown reaches first level)
            first_level = dca_levels[0]
            analysis[f'0% to -{first_level}%'] = {'count': 0, 'max_dd': 0}
            
            # Ranges between consecutive DCA levels
            for i in range(len(dca_levels) - 1):
                level1 = dca_levels[i]
                level2 = dca_levels[i + 1]
                analysis[f'-{level1}% to -{level2}%'] = {'count': 0, 'max_dd': 0}
            
            # Beyond deepest level
            deepest_level = dca_levels[-1]
            analysis[f'-{deepest_level}%+'] = {'count': 0, 'max_dd': 0}
        
        # Categorize trades
        for trade in self.trades:
            # Calculate drawdown for this trade
            drawdown_pct = abs((trade.deepest_price - trade.anchor_price) / trade.anchor_price) * 100
            
            # Skip if no meaningful drawdown
            if drawdown_pct == 0:
                continue
            
            # Find appropriate range
            range_key = None
            if dca_levels:
                first_level = dca_levels[0]
                
                if drawdown_pct < first_level:
                    # Closed before first DCA level (TP hit before drawdown reached -6%)
                    range_key = f'0% to -{first_level}%'
                else:
                    # Find which DCA level range the drawdown falls into
                    for i in range(len(dca_levels) - 1):
                        if dca_levels[i] <= drawdown_pct < dca_levels[i + 1]:
                            range_key = f'-{dca_levels[i]}% to -{dca_levels[i + 1]}%'
                            break
                    
                    # If at or beyond deepest level
                    if range_key is None:
                        range_key = f'-{dca_levels[-1]}%+'
                
                if range_key and range_key in analysis:
                    analysis[range_key]['count'] += 1
                    analysis[range_key]['max_dd'] = max(analysis[range_key]['max_dd'], drawdown_pct)
        
        return analysis
    
    def generate(self):
        """Generate the complete report."""
        self._set_column_widths()
        
        # Strategy Overview
        self._add_section_header("STRATEGY OVERVIEW")
        asset_name = self._extract_asset_name()
        self._add_key_value("Asset", asset_name)
        self._add_key_value("Period Start", self.config.start_date or "Start")
        self._add_key_value("Period End", self.config.end_date or "End")
        self._add_key_value("Strategy Type", "Mean Reversion")
        self._add_key_value("Model", "DCA (Spot)")
        dca_levels_str = ", ".join([str(abs(x)) for x in self.config.dca_levels])
        self._add_key_value("DCA Levels (%)", dca_levels_str)
        self._add_key_value("Take Profit (%)", self.config.take_profit_percent)
        self._add_key_value("Stop Loss (%)", self.config.stop_loss_percent)
        
        # Core Performance Metrics
        self._add_section_header("CORE PERFORMANCE METRICS")
        total_profit = sum(t.profit_loss for t in self.trades)
        win_count = sum(1 for t in self.trades if t.profit_loss > 0)
        loss_count = sum(1 for t in self.trades if t.profit_loss <= 0)
        win_rate = (win_count / len(self.trades) * 100) if self.trades else 0
        loss_rate = 100 - win_rate
        roi = (total_profit / self.strategy.initial_budget) * 100 if self.strategy.initial_budget > 0 else 0
        
        self._add_key_value("Total Trades", len(self.trades), 'number')
        self._add_key_value("Winning Trades", win_count, 'number')
        self._add_key_value("Win Rate (%)", win_rate / 100, 'percentage')
        self._add_key_value("Losing Trades", loss_count, 'number')
        self._add_key_value("Loss Rate (%)", loss_rate / 100, 'percentage')
        self._add_key_value("Total Net Profit ($)", total_profit, 'currency')
        self._add_key_value("Total Net Profit (%)", roi / 100, 'percentage')
        
        # Stop-Loss Metrics (if enabled)
        if self.config.stop_loss_percent > 0:
            self._add_section_header("STOP-LOSS ANALYSIS")
            stopped_out = sum(1 for t in self.trades if t.stop_loss_triggered)
            sl_loss_total = sum(t.stop_loss_loss for t in self.trades if t.stop_loss_triggered)
            sl_rate = (stopped_out / len(self.trades) * 100) if self.trades else 0
            avg_sl_loss = (sl_loss_total / stopped_out) if stopped_out > 0 else 0
            
            self._add_key_value("Stopped Out Trades", stopped_out, 'number')
            self._add_key_value("Stop-Loss Rate (%)", sl_rate / 100, 'percentage')
            self._add_key_value("Total SL Losses ($)", sl_loss_total, 'currency')
            self._add_key_value("Average SL Loss ($)", avg_sl_loss, 'currency')
        
        # Calculate average monthly return
        monthly_stats = self._calculate_monthly_stats()
        monthly_returns = [m['profit_pct'] for m in monthly_stats.values()]
        avg_monthly_return = sum(monthly_returns) / len(monthly_returns) if monthly_returns else 0
        self._add_key_value("Average Monthly Return (%)", avg_monthly_return / 100, 'percentage')
        
        # Monthly Performance Table
        self._add_section_header("MONTHLY PERFORMANCE")
        self._add_table_header(["Month", "Trade Count", "Profit ($)", "Profit (%)"])
        
        for month in sorted(monthly_stats.keys()):
            stats = monthly_stats[month]
            self._add_table_row(
                [month, stats['trades'], stats['profit'], stats['profit_pct'] / 100],
                ['text', 'number', 'currency', 'percentage']
            )
        
        # Capital Utilization & Efficiency
        self._add_section_header("CAPITAL UTILIZATION & EFFICIENCY")
        capital_stats = self._calculate_capital_stats()
        self._add_key_value("Average Capital Used (%)", capital_stats.get('avg_capital_used_pct', 0) / 100, 'percentage')
        self._add_key_value("Peak Capital Used (%)", capital_stats.get('peak_capital_used_pct', 0) / 100, 'percentage')
        self._add_key_value("Capital Buffer (Minimum Remaining %)", capital_stats.get('capital_buffer_pct', 0) / 100, 'percentage')
        self._add_key_value("Capital Efficiency", capital_stats.get('capital_efficiency', 0), 'number')
        
        # Drawdown Analysis
        self._add_section_header("DRAWDOWN ANALYSIS")
        self._add_table_header(["Drawdown Range", "Trade Count", "Max Drawdown (%)"])
        
        drawdown_analysis = self._calculate_drawdown_analysis()
        
        # Display in order (sorted by first number in range)
        for range_label in sorted(drawdown_analysis.keys(), 
                                  key=lambda x: float(x.split('%')[0].lstrip('-')) if x[0] == '-' else float(x.split('%')[0])):
            data = drawdown_analysis[range_label]
            self._add_table_row(
                [range_label, data['count'], data['max_dd'] / 100],
                ['text', 'number', 'percentage']
            )
        
        # Trade Duration Statistics
        self._add_section_header("TRADE DURATION STATISTICS")
        duration_stats = self._calculate_duration_stats()
        self._add_key_value("Average Trade Duration (Days)", duration_stats.get('avg_duration', 0), 'number')
        self._add_key_value("Longest Trade Duration (Days)", duration_stats.get('max_duration', 0), 'number')
        self._add_key_value("Trades Longer Than 30 Days", duration_stats.get('longer_than_30d', 0), 'number')
        
        # Set print settings
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_LETTER
        self.ws.page_margins.left = 0.5
        self.ws.page_margins.right = 0.5
        self.ws.page_margins.top = 0.75
        self.ws.page_margins.bottom = 0.75
        
        return self.wb
    
    def save(self, filepath: str):
        """Save the workbook to a file."""
        self.wb.save(filepath)


def run_backtest_and_generate_report(config_file: str = "strategy_config.json"):
    """
    Run backtest and generate Excel report.
    
    Args:
        config_file: Path to configuration JSON file
    """
    # Load configuration
    print("Loading configuration...")
    config = load_config(config_file)
    config.print_config()
    
    # Load data
    print(f"\nLoading {config.csv_file}...")
    df = load_and_prepare_data(config.csv_file)
    
    # Filter by date range if specified
    if config.start_date:
        start = pd.to_datetime(config.start_date)
        df = df[df['datetime'] >= start]
        print(f"Filtered to start date: {start}")
    
    if config.end_date:
        end = pd.to_datetime(config.end_date)
        df = df[df['datetime'] <= end]
        print(f"Filtered to end date: {end}")
    
    print(f"\nBacktest data: {len(df)} candles")
    print(f"Date range: {df['datetime'].iloc[0]} to {df['datetime'].iloc[-1]}")
    
    # Create strategy with all config parameters
    print(f"\nInitializing strategy...")
    strategy = DCAStrategy(
        initial_budget=config.initial_budget,
        budget_per_level=config.budget_allocation,
        dca_levels=config.dca_levels,
        take_profit_percent=config.take_profit_percent,
        stop_loss_percent=config.stop_loss_percent,
    )
    
    # Run backtest
    print(f"\nRunning backtest...")
    trades = strategy.run_backtest(df)
    
    # Print results summary
    print(f"\n{'='*70}")
    print(f"BACKTEST COMPLETED")
    print(f"{'='*70}")
    print(f"Total trades: {len(trades)}")
    if trades:
        total_pnl = sum(t.profit_loss for t in trades)
        print(f"Total P&L: ${total_pnl:,.2f}")
        print(f"ROI: {(total_pnl / strategy.initial_budget) * 100:.2f}%")
    
    # Generate Excel report
    print(f"\nGenerating Excel report...")
    generator = ExcelReportGenerator(config, strategy, trades)
    wb = generator.generate()
    
    # Generate filename with timestamp
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d-%H-%M-%S")
    csv_filename = Path(config.csv_file).stem  # e.g., "xagusd_15m_converted"
    filename = f"{timestamp}-{csv_filename}.xlsx"
    filepath = f"reports/{filename}"
    
    # Ensure reports directory exists
    Path("reports").mkdir(parents=True, exist_ok=True)
    
    generator.save(filepath)
    
    print(f"\n✓ Excel report generated successfully!")
    print(f"📊 File: {filepath}")
    
    return filepath


def main():
    if len(sys.argv) > 1:
        config_file = sys.argv[1]
    else:
        config_file = "strategy_config.json"
    
    print(f"{'='*70}")
    print(f"DCA STRATEGY - EXCEL REPORT GENERATOR")
    print(f"{'='*70}")
    print(f"Config file: {config_file}\n")
    
    try:
        filepath = run_backtest_and_generate_report(config_file)
        print(f"\n{'='*70}")
        print(f"SUCCESS")
        print(f"{'='*70}")
        print(f"Report saved to: {filepath}")
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
